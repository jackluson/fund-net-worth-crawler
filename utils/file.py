'''
Desc:
File: /file.py
File Created: Saturday, 4th March 2023 8:06:04 pm
Author: luxuemin2108@gmail.com
-----
Copyright (c) 2022 Camel Lu
'''
import os
import json

import pandas as pd
from openpyxl import load_workbook


def write_fund_json_data(data, filename, file_dir=None):
    if not file_dir:
        # cur_date = time.strftime("%Y-%m-%d", time.localtime(time.time()))
        file_dir = os.getcwd() + '/data/json/'
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
        print("目录新建成功：%s" % file_dir)
    with open(file_dir + filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.close()


def update_xlsx_file(path, df_data, sheet_name, *, index=False):
    try:
        if os.path.exists(path):
            writer = pd.ExcelWriter(path, engine='openpyxl')
            book = load_workbook(path)
            # 表名重复，删掉，重写
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            if len(book.sheetnames) == 0:
                df_data.to_excel(
                    path, sheet_name=sheet_name, index=index)
                return
            else:
                writer.book = book
            df_data.to_excel(
                writer, sheet_name=sheet_name, index=index)

            writer.save()
            writer.close()
        else:
            df_data.to_excel(
                path, sheet_name=sheet_name, index=index)
    except BaseException:
        raise BaseException('更新excel失败')
